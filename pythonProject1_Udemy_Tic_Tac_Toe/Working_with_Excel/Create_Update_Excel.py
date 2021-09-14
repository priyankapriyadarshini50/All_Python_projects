from openpyxl import Workbook
wb = Workbook('C:\\Users\\umesh\\PycharmProjects\\pythonProject1_Udemy_Tic_Tac_Toe\\Working_with_Excel')
ws = wb.active
# create a new worksheet
ws1 = wb.create_sheet("My_1st_sheet")
ws2 = wb.create_sheet()
ws2.title = 'My_2nd_sheet'
# change the tab color
ws1.sheet_properties.tabColor = "1072BA"
print(wb['My_1st_sheet'])

print(wb.sheetnames)
print(wb.path)

